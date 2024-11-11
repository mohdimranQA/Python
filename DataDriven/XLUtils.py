import openpyxl

def getRowsCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readData(file, sheetName,rowNumber,columnNumber):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return  sheet.cell(row=rowNumber,column=columnNumber).value

def writeData(file, sheetName,rowNumber,columnNumber, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNumber,column=columnNumber).value= data
    workbook.save(file)

path="/home/imran/Documents/Python/DataDriven/TestData.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name("Sheet1")
rows = sheet.max_row
cols = sheet.max_column

for r in range(1,rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r,column=c).value, end="  ")